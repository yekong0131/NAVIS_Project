import os
from django.conf import settings
from django.core.management.base import BaseCommand
from core.models import Port
from openpyxl import load_workbook  # 엑셀 처리를 위한 라이브러리


class Command(BaseCommand):
    help = "엑셀 파일에서 항구 데이터를 읽어와 DB에 저장합니다."

    def handle(self, *args, **options):
        # 엑셀 파일 경로 (파일명이 ports_data.xlsx 라고 가정)
        excel_path = os.path.join(settings.BASE_DIR, "ports_data.xlsx")

        if not os.path.exists(excel_path):
            self.stdout.write(
                self.style.ERROR(f"파일을 찾을 수 없습니다: {excel_path}")
            )
            return

        self.stdout.write(self.style.SUCCESS(f"데이터 가져오기 시작: {excel_path}"))

        try:
            # 1. 엑셀 파일 로드 (data_only=True는 수식이 아닌 값만 가져옴)
            wb = load_workbook(filename=excel_path, data_only=True)
            ws = wb.active  # 활성화된 첫 번째 시트 선택

            # 2. 헤더 처리 (첫 번째 줄)
            # 엑셀의 데이터를 [row1, row2...] 형태의 제너레이터로 가져옴
            rows = ws.iter_rows(values_only=True)

            # 첫 줄(헤더) 추출
            headers = next(rows)

            # 헤더에 공백이 있을 수 있으므로 앞뒤 공백 제거
            headers = [str(h).strip() if h else "" for h in headers]

            ports_list = []

            # 3. 데이터 루프 (두 번째 줄부터 끝까지)
            for row in rows:
                # 헤더와 현재 행의 값을 맵핑하여 딕셔너리 생성
                # 예: {'어항명': '창대항', 'Latitude': 36.406, ...}
                row_data = dict(zip(headers, row))

                # 필수 값이 없는 행(빈 행 등)은 건너뜀
                if not row_data.get("어항명") or not row_data.get("Latitude"):
                    continue

                try:
                    port = Port(
                        port_name=row_data.get("어항명"),
                        address=row_data.get("주소", ""),  # 주소가 없으면 빈 문자열
                        # 엑셀 숫자는 바로 float으로 들어오지만, 문자열일 경우를 대비해 변환
                        lat=float(row_data["Latitude"]),
                        lon=float(row_data["Longitude"]),
                        # 비고가 None일 경우 빈 문자열로 처리
                        remarks=row_data.get("비고") if row_data.get("비고") else "",
                    )
                    ports_list.append(port)

                except (ValueError, TypeError) as e:
                    self.stdout.write(
                        self.style.WARNING(
                            f"데이터 변환 오류 ({row_data.get('어항명')}): {e}"
                        )
                    )
                    continue

            # 4. DB 저장
            if ports_list:
                # 데이터 중복 방지 (선택 사항: 기존 데이터 삭제)
                # Port.objects.all().delete()

                Port.objects.bulk_create(ports_list)
                self.stdout.write(
                    self.style.SUCCESS(
                        f"총 {len(ports_list)}개의 항구 데이터 저장 완료."
                    )
                )
            else:
                self.stdout.write(self.style.WARNING("저장할 데이터가 없습니다."))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f"오류 발생: {e}"))
